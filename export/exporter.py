import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_csv_buffer(data):
    """Export to CSV format and return a string buffer."""
    if not data:
        return ""
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()

def export_json_buffer(data):
    """Export to JSON format and return a string."""
    return json.dumps({'products': data}, indent=4)

def export_excel_buffer(data):
    """Export to formatted Excel and return bytes."""
    if not data:
        return b""
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Tracking"
    
    # Headers
    headers = list(data[0].keys())
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # Data rows
    row_fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    for r_idx, row_data in enumerate(data, 2):
        for c_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = row_data[key]
            if key == "Current Price" and val is not None:
                cell.value = float(val)
                cell.number_format = '"$"#,##0.00'
            else:
                cell.value = str(val) if val is not None else ""
                
            if r_idx % 2 == 0:
                cell.fill = row_fill_alt
                
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
    # Freeze header
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
